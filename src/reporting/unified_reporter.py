"""
Unified Report Generator
========================
Author: AeroGuardian Team (Tiny Coders)
Date: 2026-01-16
Updated: 2026-02-04

Generates safety reports from FAA UAS sighting analysis:

SAFETY REPORTS (per-sighting):
- JSON: Machine-readable full data structure with telemetry metrics
- PDF: Professional single-page executive report

EVALUATION OUTPUTS (separate):
- Excel: ESRI evaluation metrics (SFS, BRR, ECC scores)

TELEMETRY ANALYSIS:
- Normalizes angles to [-180, 180] range
- Calculates GPS variance with proper coordinate filtering
- Detects anomalies using physics-based thresholds
"""

import json
import logging
import math
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import Dict, List, Any

logger = logging.getLogger("AeroGuardian.Reporter")

REPORT_TITLE = "UAS PRE-FLIGHT RISK ADVISORY"

# Import NEW 4-metric evaluation module (SFS, BRR, ECC, ESRI)
try:
    from src.evaluation.evaluate_case import (
        get_case_evaluator,
        EvaluationExcelExporter,
    )
    HAS_EVALUATION = True
except ImportError as e:
    HAS_EVALUATION = False
    logger.warning(f"New evaluation module not available: {e}")

# Try importing optional dependencies
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed - Excel export disabled")

try:
    from src.core.pdf_report_generator import PDFGenerator as PDFGenerator, HAS_REPORTLAB
except ImportError:
    HAS_REPORTLAB = False
    PDFGenerator = None
    logger.warning("PDF generator not available")

# =============================================================================
# EVALUATION TOGGLE
# Set to True to skip evaluation (faster demos, no trust score validation)
# Set to False to enable full 4-metric evaluation (SFS, BRR, ECC, ESRI)
# =============================================================================
SKIP_EVALUATION = False  # ← Change to True to disable evaluation temporarily


@dataclass
class PreFlightReport:
    """Pre-flight safety report standard."""
    
    # Header
    report_id: str
    location: str
    generated_at: str
    
    # Hazard Assessment
    hazard_level: str  # CRITICAL, HIGH, MEDIUM, LOW
    fault_probability: float
    expected_outcome: str
    
    # Root Cause Analysis
    root_cause: str
    contributing_factors: List[str]
    evidence_from_simulation: Dict[str, Any]
    
    # Prevention Recommendations
    equipment_checklist: List[str]
    environmental_checklist: List[str]
    operational_checklist: List[str]
    
    # Flight Decision
    go_nogo: str
    decision_rationale: str
    
    # Metadata
    telemetry_count: int
    flight_duration: float
    confidence: float
    
    def to_dict(self) -> Dict:
        return asdict(self)


class UnifiedReporter:
    """
    Generate PDF, Excel, and JSON reports standard.
    
    Uses existing PDFReportGenerator for PDF output.
    """
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize PDF generator if available
        self.pdf_generator = None
        if HAS_REPORTLAB and PDFGenerator:
            try:
                self.pdf_generator = PDFGenerator(self.output_dir)
                logger.info("Using PDF generator")
            except Exception as e:
                logger.warning(f"PDF generator init failed: {e}")
    
    def generate(
        self,
        incident: Dict,
        flight_config: Dict,
        telemetry: List[Dict],
        safety_analysis: Dict,
        llm_latency_ms: float = 0.0,
    ) -> Dict[str, Path]:
        """
        Generate all report formats with integrated evaluation.
        
        Output Structure:
        - {incident_id}/report/report.json, report.xlsx, report.pdf
        - {incident_id}/evaluation/evaluation.json, evaluation.xlsx
        
        Args:
            incident: FAA incident data
            flight_config: LLM-generated PX4 configuration
            telemetry: Simulation telemetry data
            safety_analysis: Generated safety analysis
            llm_latency_ms: LLM call latency for logging
            
        Returns:
            Dict with paths to generated files
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        incident_id = incident.get('report_id', 'unknown')
        incident_dir = self.output_dir / f"{incident_id}_{timestamp}"
        incident_dir.mkdir(exist_ok=True)
        
        # Create subfolders
        report_dir = incident_dir / "report"
        report_dir.mkdir(exist_ok=True)
        
        generated_dir = incident_dir / "generated"
        generated_dir.mkdir(exist_ok=True)

        # Build unified report data
        report_data = self._build_report_data(incident, flight_config, telemetry, safety_analysis)
        
        # =========================================================================
        # EVALUATION (4-metric system: SFS, BRR, ECC, ESRI)
        # Set SKIP_EVALUATION=True at top of file to disable for faster demos
        # =========================================================================
        
        # Run NEW 4-metric evaluation (SFS, BRR, ECC, ESRI)
        evaluation_result = None
        if SKIP_EVALUATION:
            logger.info("⏭️ Evaluation SKIPPED (SKIP_EVALUATION=True)")
        elif HAS_EVALUATION:
            try:
                evaluator = get_case_evaluator()
                # IMPORTANT: Pass None for telemetry_stats so BRR evaluator
                # computes its own stats from raw telemetry data.
                # The report_data telemetry_summary has different key names
                # that don't match what BRR expects.
                evaluation_result = evaluator.evaluate(
                    faa_report=incident,
                    px4_config=flight_config,
                    telemetry=telemetry,
                    safety_report=safety_analysis,
                    telemetry_stats=None,  # Let BRR compute from raw telemetry
                )
                
                report_data["evaluation"] = evaluation_result.to_dict()
                logger.info(
                    f"Evaluation: ESRI={evaluation_result.esri * 100:.1f}% "
                    f"(SFS={evaluation_result.sfs * 100:.0f}%, BRR={evaluation_result.brr * 100:.0f}%, "
                    f"ECC={evaluation_result.ecc * 100:.0f}%) [{evaluation_result.consistency_level}]"
                )
            except Exception as e:
                logger.warning(f"Evaluation failed: {e}")
        
        paths = {}
        
        # =================================================================
        # GENERATED FILES (in generated/ subfolder)
        # Full LLM configuration and telemetry for traceability
        # =================================================================
        
        # Save full LLM configuration output
        config_path = generated_dir / "full_configuration_output_from_llm.json"
        self._save_full_config(flight_config, incident, config_path)
        paths["full_config"] = config_path
        
        # Save full telemetry output
        telemetry_path = generated_dir / "full_telemetry_of_each_flight.json"
        self._save_full_telemetry(telemetry, incident, telemetry_path)
        paths["full_telemetry"] = telemetry_path
        
        # =================================================================
        # REPORT FILES (in report/ subfolder)
        # =================================================================
        
        # Generate JSON
        json_path = report_dir / "report.json"
        self._generate_json(report_data, telemetry, json_path)
        paths["json"] = json_path
        
        # Generate PDF (if reportlab available)
        if self.pdf_generator:
            pdf_path = report_dir / "report.pdf"
            self._generate_pdf(report_data, pdf_path)
            paths["pdf"] = pdf_path
        
        # =================================================================
        # EVALUATION FILES (in evaluation/ subfolder)
        # Research-grade evaluation export is currently ENABLED
        # =================================================================
        
        if evaluation_result and HAS_EVALUATION:
            try:
                eval_dir = incident_dir / "evaluation"
                eval_dir.mkdir(exist_ok=True)
                
                # Import regulatory standards for inclusion in evaluation
                try:
                    from src.evaluation.regulatory_standards import get_regulatory_summary
                    regulatory_data = get_regulatory_summary()
                except ImportError:
                    regulatory_data = {}
                
                # Export evaluation JSON with regulatory grounding
                eval_json_path = eval_dir / "evaluation.json"
                import json
                
                # Combine evaluation results with regulatory standards
                eval_data = evaluation_result.to_dict()
                eval_data.update(regulatory_data)  # Include FAA/DO standards
                
                with open(eval_json_path, "w", encoding="utf-8") as f:
                    json.dump(eval_data, f, indent=2)
                paths["evaluation_json"] = eval_json_path
                
                # Export evaluation Excel
                eval_exporter = EvaluationExcelExporter(eval_dir)
                eval_excel_path = eval_exporter.export_incident(evaluation_result)
                if eval_excel_path:
                    paths["evaluation_excel"] = eval_excel_path
                    
                logger.info(f"Evaluation exported: ESRI={evaluation_result.esri:.3f}")
            except Exception as e:
                logger.warning(f"Evaluation export failed: {e}")
        
        logger.info(f"Reports saved to: {incident_dir}")
        paths["report_dir"] = incident_dir
        return paths
    
    def _extract_assumptions(self, flight_config: Dict) -> List[str]:
        """
        Extract assumptions made during scenario generation.
        
        SAFETY RATIONALE: Every assumption reduces confidence.
        We must track what was inferred vs. what was explicit in source.
        """
        assumptions = []
        
        # Check for default values that indicate missing data
        reasoning = flight_config.get("reasoning", "")
        
        if "default" in reasoning.lower():
            assumptions.append("Default values used where FAA report lacked specifics")
        
        if "assumed" in reasoning.lower() or "infer" in reasoning.lower():
            assumptions.append("Failure mode inferred from behavioral description")
        
        if "not specified" in reasoning.lower():
            assumptions.append("Some parameters not specified in source report")
        
        # Check environment assumptions
        env = flight_config.get("environment", {})
        if env.get("weather") in ["not_specified", "unknown", None, ""]:
            assumptions.append("Weather conditions unknown, using defaults")
        
        # Note: Altitude clamping (previously 120m) was removed to support full high-altitude modeling
        return assumptions
    
    def _build_report_data(
        self,
        incident: Dict,
        flight_config: Dict,
        telemetry: List[Dict],
        safety_analysis: Dict,
    ) -> Dict:
        """Build report data structure with 6-section Pre-Flight Safety Report format."""
        
        # Validate required inputs
        if not incident:
            raise ValueError("incident data required")
        if not safety_analysis:
            raise ValueError("safety_analysis data required")
        
        # Extract telemetry statistics
        telemetry_stats = self._analyze_telemetry(telemetry) if telemetry else {}

        # Extract causal fields if available from upstream analysis
        primary_failure_subsystem = safety_analysis.get("primary_failure_subsystem", "undetermined")
        causal_chain = safety_analysis.get("causal_chain", [])
        if isinstance(causal_chain, str):
            causal_chain = [c.strip() for c in causal_chain.split("->") if c.strip()]
        subsystem_confidence = safety_analysis.get("subsystem_confidence", safety_analysis.get("confidence", 0.0))
        
        # Extract go/no-go verdict
        verdict = safety_analysis.get("verdict", "REVIEW")
        if "NO-GO" in verdict.upper() or "NO_GO" in verdict.upper():
            verdict = "NO-GO"
        elif "CAUTION" in verdict.upper():
            verdict = "CAUTION"
        else:
            verdict = "GO" if verdict.upper() == "GO" else verdict
        
        # Build the NEW 3-SECTION report structure
        return {
            "report_type": REPORT_TITLE,
            "version": "1.0",
            "generated_at": datetime.now().isoformat(),
            
            # =========================================================
            # MANDATORY DISCLAIMER (SAFETY-CRITICAL)
            # =========================================================
            "disclaimer": {
                "status": "DECISION SUPPORT TOOL - NOT A SAFETY CERTIFICATION",
                "data_source": "FAA UAS sighting data (observational, non-investigative)",
                "methodology": "LLM-inferred failure scenarios (hypothesis, not diagnosis)",
                "simulation": "PX4 SITL with X500 quadrotor (proxy approximation, not reconstruction)",
                "warning": "Do NOT use this report as sole justification for flight operations. All operational decisions remain the responsibility of the operator.",
            },
            
            # =========================================================
            # INCIDENT SOURCE (Input Context)
            # =========================================================
            "incident_source": {
                "original_faa_narrative": incident.get("description", incident.get("summary", "")),
                "report_id": incident.get("report_id", "Unknown"),
                "date_time": incident.get("date", ""),
                "location": f"{incident.get('city', 'Unknown')}, {incident.get('state', 'Unknown')}",
            },
            
            # =========================================================
            # SECTION 1: SAFETY LEVEL & CAUSE
            # What went wrong and how dangerous it is
            # =========================================================
            "section_1_safety_level_and_cause": {
                "title": "1. SAFETY LEVEL, HAZARD, AND ROOT CAUSE",
                "safety_level": safety_analysis.get("safety_level", safety_analysis.get("hazard_level", "UNKNOWN")),
                "primary_hazard": safety_analysis.get("primary_hazard", safety_analysis.get("hazard_type", "Unknown hazard")),
                "root_cause_subsystem": primary_failure_subsystem,
                "causal_chain": causal_chain,
                "root_cause_confidence": subsystem_confidence,
                "observed_effect": safety_analysis.get("observed_effect", "No specific effect observed"),
            },
            
            # =========================================================
            # SECTION 2: DESIGN CONSTRAINTS & RECOMMENDATIONS
            # What MUST be constrained or changed before flight
            # =========================================================
            "section_2_design_constraints_and_recommendations": {
                "title": "2. AIRCRAFT SYSTEM DESIGN CONSTRAINTS AND MITIGATION RECOMMENDATIONS",
                "scope_note": "Design constraints apply to aircraft system configuration and operating envelope. Recommendations target operator actions and engineering mitigations.",
                "design_constraints": safety_analysis.get("design_constraints", []),
                "recommendations": safety_analysis.get("recommendations", safety_analysis.get("safety_recommendations", [])),
            },
            
            # =========================================================
            # SECTION 3: EXPLANATION (WHY)
            # Why the system reached this conclusion - THE NOVELTY
            # =========================================================
            "section_3_explanation": {
                "title": "3. EVIDENCE TRACEABILITY AND OPERATOR ACTION RATIONALE",
                "reasoning": safety_analysis.get("explanation", safety_analysis.get("conclusion", "Analysis based on FAA incident and simulation data.")),
            },
            
            # =========================================================
            # FINAL VERDICT
            # =========================================================
            "verdict": {
                "decision": verdict,
                "go_nogo": verdict,
            },

            # Structured causal fields for downstream auditing/evaluation
            "primary_hazard": safety_analysis.get("primary_hazard", safety_analysis.get("hazard_type", "Unknown hazard")),
            "primary_failure_subsystem": primary_failure_subsystem,
            "causal_analysis": {
                "primary_failure_subsystem": primary_failure_subsystem,
                "causal_chain": causal_chain,
                "confidence": subsystem_confidence,
            },
            
            # =========================================================
            # SUPPORTING DATA (Telemetry & Config)
            # =========================================================
            "supporting_data": {
                "simulation_config": {
                    "waypoints_count": len(flight_config.get("waypoints", [])),
                    "fault_type": flight_config.get("fault_injection", {}).get("fault_type", "None"),
                    "airframe_model": flight_config.get("uav_model", flight_config.get("vehicle_type", "x500_quadcopter")),
                    "altitude_m": flight_config.get("mission", {}).get("takeoff_altitude_m", 0),
                    "speed_ms": flight_config.get("speed_m_s", 5.0),
                },
                "telemetry_summary": {
                    "data_points": len(telemetry) if telemetry else 0,
                    "duration_sec": telemetry_stats.get("flight_duration_s", 0),
                    "max_altitude_m": telemetry_stats.get("max_altitude_m", 0),
                    "max_roll_deg": telemetry_stats.get("max_roll_deg", 0),
                },
            },
            
            # Legacy compatibility fields
            "executive_summary": {
                "verdict": verdict,
                "hazard_level": safety_analysis.get("safety_level", safety_analysis.get("hazard_level", "UNKNOWN")),
            },
            "incident": {
                "id": incident.get("report_id", "Unknown"),
                "location": f"{incident.get('city', 'Unknown')}, {incident.get('state', 'Unknown')}",
            },
            "flight_config": flight_config if flight_config else {},
            "telemetry_summary": {
                "count": len(telemetry) if telemetry else 0,
                "statistics": telemetry_stats,
            },
        }
    
    def _analyze_telemetry(self, telemetry: List[Dict]) -> Dict:
        """Extract statistics from telemetry for report."""
        if not telemetry:
            return {"error": "No telemetry data"}
        
        # Helper to normalize angles to ±180° range
        # This prevents misleading values like 1070° when drone rotates multiple times
        def normalize_angle(angle_deg: float) -> float:
            """Normalize angle to [-180, 180] range."""
            while angle_deg > 180:
                angle_deg -= 360
            while angle_deg < -180:
                angle_deg += 360
            return angle_deg
        
        alts = [t.get("alt", 0) for t in telemetry]

        # Robust speed extraction: prefer explicit horizontal velocity, otherwise
        # estimate from GPS deltas when velocity channels are unavailable.
        speeds = []
        prev = None
        for t in telemetry:
            spd = None

            vx = t.get("vx")
            vy = t.get("vy")
            if vx is not None and vy is not None:
                spd = (vx**2 + vy**2) ** 0.5
            else:
                vel_n = t.get("vel_n_m_s")
                vel_e = t.get("vel_e_m_s")
                if vel_n is not None and vel_e is not None:
                    spd = (vel_n**2 + vel_e**2) ** 0.5

            if spd is None and prev is not None:
                dt = t.get("timestamp", 0) - prev.get("timestamp", 0)
                if dt > 0:
                    lat1, lon1 = prev.get("lat", 0), prev.get("lon", 0)
                    lat2, lon2 = t.get("lat", 0), t.get("lon", 0)
                    if lat1 and lon1 and lat2 and lon2:
                        dlat = (lat2 - lat1) * 111000
                        dlon = (lon2 - lon1) * 111000 * abs(math.cos(math.radians(lat1)))
                        spd = math.sqrt(dlat**2 + dlon**2) / dt

            speeds.append(float(spd) if spd is not None else 0.0)
            prev = t

        # Roll/pitch may arrive either in radians or degrees depending on source.
        def to_degrees_if_needed(angle_value: float) -> float:
            return angle_value * 57.2958 if abs(angle_value) <= (2 * math.pi + 0.01) else angle_value

        pitches = [normalize_angle(to_degrees_if_needed(t.get("pitch", 0))) for t in telemetry]
        rolls = [normalize_angle(to_degrees_if_needed(t.get("roll", 0))) for t in telemetry]
        
        # Calculate GPS variance (position drift) - max distance from start position
        # CRITICAL: Only use telemetry points with valid GPS coordinates (non-zero)
        gps_variance = 0.0
        if len(telemetry) > 1:
            # Filter for valid GPS coordinates (exclude zeros which indicate missing data)
            valid_coords = [
                (t.get("lat", 0), t.get("lon", 0)) 
                for t in telemetry 
                if t.get("lat", 0) != 0 and t.get("lon", 0) != 0
            ]
            
            if len(valid_coords) > 1:
                lats = [c[0] for c in valid_coords]
                lons = [c[1] for c in valid_coords]
                start_lat, start_lon = lats[0], lons[0]
                
                # Calculate max distance from start (in meters)
                max_drift = 0.0
                for lat, lon in valid_coords:
                    dlat = (lat - start_lat) * 111000  # degrees to meters
                    dlon = (lon - start_lon) * 111000 * abs(math.cos(math.radians(start_lat)))
                    drift = math.sqrt(dlat**2 + dlon**2)
                    max_drift = max(max_drift, drift)
                gps_variance = max_drift
        
        # Calculate roll oscillation frequency (zero crossings / duration)
        roll_oscillation_freq = 0.0
        if len(rolls) > 2:
            zero_crossings = sum(1 for i in range(1, len(rolls)) if rolls[i-1] * rolls[i] < 0)
            duration = telemetry[-1].get("timestamp", 0) - telemetry[0].get("timestamp", 0)
            if duration > 0:
                roll_oscillation_freq = zero_crossings / duration
        
        # Calculate altitude deviation from target
        target_alt = max(alts) if alts else 0
        altitude_deviation = sum(abs(a - target_alt) for a in alts) / len(alts) if alts else 0
        
        # Detect failsafe events (expanded to include attitude instabilities)
        failsafe_events = []
        max_roll_abs = max(abs(r) for r in rolls) if rolls else 0
        max_pitch_abs = max(abs(p) for p in pitches) if pitches else 0
        
        # Check for motor failure pattern (excessive roll)
        if max_roll_abs > 30:  # 30° roll is abnormal for multirotors
            failsafe_events.append("MOTOR_FAILURE_PATTERN")
        
        # Check for control anomaly (excessive pitch)
        if max_pitch_abs > 20:  # 20° pitch deviation
            failsafe_events.append("CONTROL_ANOMALY")
        
        # Check for altitude-based events
        for i in range(1, len(telemetry)):
            alt_drop = telemetry[i-1].get("alt", 0) - telemetry[i].get("alt", 0)
            if alt_drop > 5:  # Sudden 5m drop
                failsafe_events.append("ALTITUDE_DROP")
            if abs(telemetry[i].get("vz", 0)) > 3:  # Fast vertical speed
                failsafe_events.append("FAST_DESCENT")
        
        # Check for position drift (GPS anomaly)
        if gps_variance > 50:  # 50m drift indicates significant issue
            failsafe_events.append("POSITION_DRIFT_CRITICAL")
        elif gps_variance > 10:
            failsafe_events.append("POSITION_DRIFT_HIGH")
        
        # Remove duplicates
        failsafe_events = list(set(failsafe_events))
        
        return {
            "flight_duration_s": telemetry[-1].get("timestamp", 0) - telemetry[0].get("timestamp", 0),
            "max_altitude_m": max(alts) if alts else 0,
            "avg_altitude_m": sum(alts) / len(alts) if alts else 0,
            "max_speed_mps": max(speeds) if speeds else 0,
            "avg_speed_mps": sum(speeds) / len(speeds) if speeds else 0,
            "max_roll_deg": round(max(abs(r) for r in rolls), 1) if rolls else 0,
            "max_pitch_deg": round(max(abs(p) for p in pitches), 1) if pitches else 0,
            "data_points": len(telemetry),
            
            # LLM-ready telemetry summary format
            "flight_summary": {
                "max_pitch": round(max(abs(p) for p in pitches), 1) if pitches else 0,
                "max_roll": round(max(abs(r) for r in rolls), 1) if rolls else 0,
                "roll_oscillation_freq": round(roll_oscillation_freq, 2),
                "gps_variance": round(gps_variance, 1),
                "altitude_deviation": round(altitude_deviation, 1),
                "control_saturation_time": 0.0,  # Would need throttle data
                "failsafe_events": failsafe_events or ["NONE"],
            },
            "expected_behavior": {
                "nominal_pitch_range": [-15, 15],
                "nominal_roll_range": [-30, 30],
                "gps_variance_max": 1.0,
                "max_roll_oscillation": 1.0,
                "altitude_tolerance": 5.0,
            },
        }
    
    def _extract_factors(self, safety_analysis: Dict) -> List[str]:
        """Extract contributing factors from safety analysis."""
        factors = []
        
        key_insight = safety_analysis.get("key_insight", "")
        if key_insight:
            # Split key insight into factors
            if ";" in key_insight:
                factors = [f.strip() for f in key_insight.split(";")]
            elif "," in key_insight:
                factors = [f.strip() for f in key_insight.split(",")[:3]]
            else:
                factors = [key_insight]
        
        return factors[:5]  # Max 5 factors
    
    def _save_full_config(self, flight_config: Dict, incident: Dict, path: Path):
        """
        Save full LLM-generated configuration output with P0-P2 research enhancements.
        
        This preserves the complete configuration with all parameters,
        uncertainty scores, evidence traceability, and proxy modeling info.
        """
        # Extract P0-P2 fields from flight_config if present
        uncertainty_score = flight_config.get("uncertainty_score", 0.5)
        fault_injection_supported = flight_config.get("fault_injection_supported", True)
        narrative_facts = flight_config.get("narrative_facts", {})
        inferred_parameters = flight_config.get("inferred_parameters", {})
        proxy_modeling = flight_config.get("proxy_modeling", {})
        evidence_map = flight_config.get("evidence_map", {})
        reconstruction_level = flight_config.get("reconstruction_level", "proxy_simulation")
        
        # Check for fault alignment issue (P0)
        injection_marker = flight_config.get("px4_commands", {}).get("fault", "unknown")
        fault_type = flight_config.get("fault_injection", {}).get("fault_type", "unknown")
        fault_alignment_warning = None
        if (not fault_injection_supported) and fault_type not in ["none", "unknown"]:
            fault_alignment_warning = (
                f"UNSUPPORTED: fault_type='{fault_type}' but fault_injection_supported=False "
                f"(injection_marker='{injection_marker}'). Using behavioral emulation."
            )
        
        # Check for fault injection execution status
        fault_injection_status = flight_config.get("fault_injection_status", {})
        fault_injection_mode = fault_injection_status.get("mode", "unknown")
        fault_injection_success = fault_injection_status.get("success", True)
        fault_injection_warning = None
        if fault_injection_mode == "fallback":
            fault_injection_warning = "FALLBACK: Fault injection failed (timeout or error). Using approximate behavioral emulation instead of actual sensor/actuator injection."
        elif fault_injection_mode == "emulated" and not fault_injection_success:
            fault_injection_warning = "EMULATED: Native PX4 fault injection unavailable. Using behavioral simulation."
        
        full_output = {
            "metadata": {
                "file_type": "full_configuration_output_from_llm",
                "generated_at": datetime.now().isoformat(),
                "report_id": incident.get("report_id", "unknown"),
                "incident_location": f"{incident.get('city', 'Unknown')}, {incident.get('state', 'Unknown')}",
                "description": "Complete LLM-generated PX4 simulation configuration",
                # P2: Reconstruction metadata
                "reconstruction_level": reconstruction_level,
                "reconstruction_level_description": {
                    "proxy_simulation": "Simulating with different platform than source aircraft",
                    "partial_match": "Same aircraft class with some inferred parameters",
                    "behavioral_class": "Simulating general failure behavior class only",
                }.get(reconstruction_level, "Unknown"),
            },
            "source_incident": {
                "id": incident.get("report_id", ""),
                "city": incident.get("city", ""),
                "state": incident.get("state", ""),
                "date": incident.get("date", ""),
                "summary": incident.get("summary", incident.get("description", "")),
                "incident_type": incident.get("incident_type", ""),
            },
            # P0: Uncertainty and alignment
            "uncertainty_analysis": {
                "uncertainty_score": uncertainty_score,
                "uncertainty_interpretation": "LOW" if uncertainty_score < 0.4 else "MEDIUM" if uncertainty_score < 0.6 else "HIGH",
                "fault_injection_supported": fault_injection_supported,
                "fault_alignment_warning": fault_alignment_warning,
                "fault_injection_status": {
                    "mode": fault_injection_mode,
                    "success": fault_injection_success,
                    "warning": fault_injection_warning,
                },
            },
            # P0: Narrative facts vs inferred parameters
            "narrative_vs_inferred": {
                "narrative_facts": narrative_facts or {
                    "note": "Extracted directly from FAA report text",
                    "location_stated": f"{incident.get('city', '')}, {incident.get('state', '')}",
                },
                "inferred_parameters": inferred_parameters or {
                    "note": "LLM-inferred values not explicitly stated in report",
                },
            },
            # P1: Proxy modeling information
            "proxy_simulation": proxy_modeling or {
                "source_aircraft_class": "unknown",
                "simulation_platform": "x500_quadcopter",
                "platform_substitution": True,
                "substitution_reason": "PX4 SITL default quadrotor model",
            },
            # P2: Evidence traceability
            "evidence_traceability": {
                "map": evidence_map or {},
                "legend": {
                    "FAA_NARRATIVE": "Extracted from FAA sighting report text",
                    "LLM_INFERENCE": "Inferred by LLM from context",
                    "LLM_DEFAULT": "LLM used default value (not in report)",
                    "LLM_MAPPING": "LLM mapped to PX4 command",
                    "LLM_GENERATED": "LLM generated (waypoints, etc.)",
                    "GEOCODER_API": "Resolved via geocoding API",
                    "NOT_SUPPORTED": "Cannot be represented in simulation",
                },
            },
            "llm_configuration": flight_config,
            "parameter_count": self._count_params(flight_config),
        }
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(full_output, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"  Full config: {path.name} ({full_output['parameter_count']} parameters)")
    
    def _save_full_telemetry(self, telemetry: List[Dict], incident: Dict, path: Path):
        """
        Save full telemetry output from virtual flight.
        
        This preserves the complete flight telemetry data with all
        sensor readings for detailed analysis and research.
        """
        # Calculate flight statistics
        duration = 0
        max_alt = 0
        if telemetry:
            timestamps = [t.get("timestamp", 0) for t in telemetry]
            altitudes = [t.get("alt", t.get("altitude", 0)) for t in telemetry]
            duration = max(timestamps) - min(timestamps) if timestamps else 0
            max_alt = max(altitudes) if altitudes else 0
        
        full_output = {
            "metadata": {
                "file_type": "full_telemetry_of_each_flight",
                "generated_at": datetime.now().isoformat(),
                "report_id": incident.get("report_id", "unknown"),
                "description": "Complete flight telemetry from PX4 simulation",
            },
            "flight_summary": {
                "total_data_points": len(telemetry),
                "flight_duration_sec": round(duration, 2),
                "max_altitude_m": round(max_alt, 2),
                "sampling_rate_hz": len(telemetry) / duration if duration > 0 else 0,
            },
            "telemetry": telemetry,
        }
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(full_output, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"  Full telemetry: {path.name} ({len(telemetry)} data points)")
    
    def _count_params(self, config: Dict, prefix: str = "") -> int:
        """Count parameters in nested config dictionary."""
        count = 0
        for key, value in config.items():
            if isinstance(value, dict):
                count += self._count_params(value, f"{prefix}{key}.")
            elif isinstance(value, list):
                count += len(value) if value else 1
            else:
                count += 1
        return count
    
    def _generate_json(self, report_data: Dict, telemetry: List[Dict], path: Path):
        """Generate JSON report - NO raw telemetry, only summary for smaller file size."""
        # Don't include raw telemetry array - it's already summarized in telemetry_summary
        # This reduces JSON file size by ~60%
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"  JSON report: {path.name}")
    
    def _generate_excel(self, report_data: Dict, telemetry: List[Dict], path: Path, evaluation_result=None):
        """Generate Excel report with multiple sheets including Evaluation."""
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Add summary data using 6-section structure
        s1 = report_data.get("section_1_incident_source", {})
        s2 = report_data.get("section_2_scenario_translation", {})
        s4 = report_data.get("section_4_telemetry_summary", {})
        s5 = report_data.get("section_5_safety_analysis", {})
        s6 = report_data.get("section_6_conclusion", {})
        
        summary_data = [
            (REPORT_TITLE, ""),
            ("", ""),
            # ("SECTION 1: INCIDENT SOURCE", ""),
            ("Report ID", s1.get("report_id", report_data.get("incident", {}).get("id", "Unknown"))),
            ("Location", s1.get("location", report_data.get("incident", {}).get("location", "Unknown"))),
            ("Date/Time", s1.get("date_time", "Unknown")),
            ("", ""),
            ("SECTION 2: SCENARIO TRANSLATION", ""),
            ("Hazard Type", s2.get("hazard_type", "Unknown")),
            ("Altitude", f"{s2.get('altitude', 0):.0f}m"),
            ("", ""),
            ("SECTION 4: TELEMETRY SUMMARY", ""),
            ("Duration", f"{s4.get('duration_sec', 0)} seconds"),
            ("Data Points", report_data.get("telemetry_summary", {}).get("count", 0)),
            ("", ""),
            ("SECTION 5: SAFETY ANALYSIS", ""),
            ("Risk Level", s5.get("risk_level", "UNKNOWN")),
            ("", ""),
            ("SECTION 6: CONCLUSION", ""),
            ("Verdict", s6.get("verdict", "REVIEW")),
            ("Recommendation", s6.get("go_nogo_recommendation", "")),
            ("Generated", report_data.get("generated_at", "")),
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)
            
            if label.startswith("SECTION") or label == REPORT_TITLE:
                ws_summary.cell(row=row_idx, column=1).fill = header_fill
                ws_summary.cell(row=row_idx, column=1).font = header_font
        
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 60
        
        # Sheet 2: Telemetry
        ws_telemetry = wb.create_sheet("Telemetry")
        
        if telemetry:
            headers = ["timestamp", "lat", "lon", "alt", "vx", "vy", "vz", "roll", "pitch", "yaw"]
            for col, header in enumerate(headers, 1):
                cell = ws_telemetry.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, t in enumerate(telemetry[:1000], 2):  # Limit to 1000 rows
                for col, header in enumerate(headers, 1):
                    ws_telemetry.cell(row=row_idx, column=col, value=t.get(header, 0))
        
        # Sheet 3: Checklist (using new 6-section structure)
        ws_checklist = wb.create_sheet("Pre-Flight Checklist")
        
        # Get safety recommendations from section 5
        safety_recs = s5.get("safety_recommendations", [])
        if isinstance(safety_recs, str):
            safety_recs = [r.strip() for r in safety_recs.split("|") if r.strip()]
        
        # Get hazard indicators from section 4
        hazards = s4.get("hazard_indicators", [])
        if isinstance(hazards, str):
            hazards = [h.strip() for h in hazards.split("|") if h.strip()]
        
        checklist_rows = [
            ("PRE-FLIGHT SAFETY CHECKLIST", ""),
            ("", ""),
            ("SAFETY RECOMMENDATIONS", ""),
        ] + [(f"☐ {item}", "") for item in safety_recs] + [
            ("", ""),
            ("HAZARD INDICATORS TO MONITOR", ""),
        ] + [(f">>>>>  {item}", "") for item in hazards] + [
            ("", ""),
            ("VERDICT", s6.get("verdict", "REVIEW")),
            ("GO/NO-GO", s6.get("go_nogo_recommendation", "")),
        ]
        
        for row_idx, (item, val) in enumerate(checklist_rows, 1):
            ws_checklist.cell(row=row_idx, column=1, value=item)
            ws_checklist.cell(row=row_idx, column=2, value=val)
            if item in ["PRE-FLIGHT SAFETY CHECKLIST", "SAFETY RECOMMENDATIONS", "HAZARD INDICATORS TO MONITOR", "VERDICT"]:
                ws_checklist.cell(row=row_idx, column=1).font = Font(bold=True)
        
        ws_checklist.column_dimensions['A'].width = 60
        ws_checklist.column_dimensions['B'].width = 30
        
        # Sheet 4: Evaluation (Research Metrics)
        # if evaluation_result:
        #     ws_eval = wb.create_sheet("Evaluation")
            
        #     eval_rows = [
        #         ("RESEARCH EVALUATION METRICS", ""),
        #         ("", ""),
        #         ("Incident ID", evaluation_result.incident_id),
        #         ("Evaluation Timestamp", evaluation_result.evaluation_timestamp),
        #         ("Overall Score", f"{evaluation_result.overall_score}/100"),
        #         ("Grade", evaluation_result.grade),
        #         ("", ""),
        #         ("INPUT FIDELITY", ""),
        #         ("Extraction Score", f"{evaluation_result.input_fidelity.extraction_score:.1%}"),
        #         ("Fault Type Extracted", ">>>>> " if evaluation_result.input_fidelity.fault_type_extracted else "✗"),
        #         ("Location Extracted", ">>>>> " if evaluation_result.input_fidelity.location_extracted else "✗"),
        #         ("Waypoints Generated", ">>>>> " if evaluation_result.input_fidelity.waypoints_generated else "✗"),
        #         ("LLM Latency (ms)", f"{evaluation_result.input_fidelity.llm_response_time_ms:.0f}"),
        #         ("", ""),
        #         ("SIMULATION VALIDITY", ""),
        #         ("Behavior Match Score", f"{evaluation_result.simulation_validity.behavior_match_score:.1%}"),
        #         ("Semantic Similarity", f"{evaluation_result.simulation_validity.semantic_similarity:.2f}"),
        #         ("Telemetry Points", evaluation_result.simulation_validity.telemetry_points_count),
        #         ("Simulation Completed", ">>>>> " if evaluation_result.simulation_validity.simulation_completed else "✗"),
        #         ("Fault Manifested", ">>>>> " if evaluation_result.simulation_validity.fault_manifested else "✗"),
        #         ("", ""),
        #         ("OUTPUT UTILITY", ""),
        #         ("Prevention Score", f"{evaluation_result.output_utility.prevention_score}/100"),
        #         ("Hazard Level", evaluation_result.output_utility.hazard_level),
        #         ("Go/No-Go Verdict", evaluation_result.output_utility.go_nogo_verdict),
        #         ("Specificity Score", f"{evaluation_result.output_utility.specificity_score:.1f}/5"),
        #         ("Actionability Score", f"{evaluation_result.output_utility.actionability_score:.1f}/5"),
        #         ("Checklist Items", evaluation_result.output_utility.checklist_count),
        #         ("", ""),
        #         ("RESEARCH TARGETS", ""),
        #     ]
            
        #     # Add targets met
        #     for target_name, met in evaluation_result.targets_met.items():
        #         eval_rows.append((target_name.replace("_", " ").title(), ">>>>> " if met else "✗"))
            
        #     for row_idx, (label, value) in enumerate(eval_rows, 1):
        #         cell_label = ws_eval.cell(row=row_idx, column=1, value=label)
        #         ws_eval.cell(row=row_idx, column=2, value=value)
                
        #         if label in ["RESEARCH EVALUATION METRICS", "INPUT FIDELITY", "SIMULATION VALIDITY", "OUTPUT UTILITY", "RESEARCH TARGETS"]:
        #             cell_label.fill = header_fill
        #             cell_label.font = header_font
            
        #     ws_eval.column_dimensions['A'].width = 30
        #     ws_eval.column_dimensions['B'].width = 25
        
        wb.save(path)
        logger.info(f"  Excel report: {path.name}")
    
    def _generate_pdf(self, report_data: Dict, path: Path):
        """Generate PDF report using existing PDFReportGenerator."""
        try:
            self.pdf_generator.generate(report_data, path)
            logger.info(f"  PDF report: {path.name}")
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")


# Convenience function
def generate_reports(
    output_dir: Path,
    incident: Dict,
    flight_config: Dict,
    telemetry: List[Dict],
    safety_analysis: Dict,
) -> Dict[str, Path]:
    """Convenience function to generate all reports."""
    reporter = UnifiedReporter(output_dir)
    return reporter.generate(incident, flight_config, telemetry, safety_analysis)
