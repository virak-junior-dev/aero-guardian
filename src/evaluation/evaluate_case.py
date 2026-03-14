"""
Case Evaluator
==============
Author: AeroGuardian Team (Tiny Coders)
Date: 2026-01-21
Updated: 2026-02-05

Unified evaluator that orchestrates all 4 research-grade metrics:
- CCR: Constraint Correctness Rate (LLM translation accuracy)
- BRR: Behavior Reproduction Rate (telemetry anomaly detection)
- ECC: Evidence-Conclusion Consistency (claim grounding)
- ESRI: Executable Safety Reliability Index (CCR/SFS × BRR × ECC)

Produces per-incident evaluation reports with trust level assessment.

SAFETY GUARDRAILS:
- ESRI is capped at 0.85 maximum (FAA source is non-authoritative)
- Confidence ceilings enforced at each evaluation stage
- All assumptions logged and penalized
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from dataclasses import dataclass, field

logger = logging.getLogger("AeroGuardian.Evaluator.Case")

# Import metric components
from .scenario_fidelity import ScenarioFidelityScorer
from .behavior_validation import BehaviorValidator
from .evidence_consistency import EvidenceConsistencyChecker
from .esri import ESRICalculator
from .subsystem_analysis import SubsystemCausalAnalyzer
from .constraint_correctness import ConstraintCorrectnessEvaluator
from .uncertainty_robustness import UncertaintyRobustnessEvaluator

# =============================================================================
# CONFIDENCE CEILING CONSTANTS
# These are HARD LIMITS that CANNOT be exceeded regardless of component scores
# =============================================================================

# SAFETY RATIONALE: FAA sighting reports are observational and non-authoritative.
# PX4 simulation is a proxy model, not a reconstruction of the actual incident.
# These ceilings prevent overconfidence in system outputs.
ESRI_ABSOLUTE_CEILING = 0.85  # Maximum possible ESRI even with perfect components
SFS_CEILING = 0.80            # Legacy label; now applied to CCR as primary LLM translation score
BRR_CEILING = 0.95            # Physics simulation accuracy
ECC_CEILING = 1.0             # Evidence matching (can be perfect if well-supported)


@dataclass
class CaseEvaluationResult:
    """Complete evaluation result for a single incident case."""
    
    incident_id: str
    evaluation_timestamp: str
    
    # Metric results
    ccr: float
    sfs: float
    brr: float
    ecc: float
    agi: float
    urs: float
    ees: float
    esri: float
    
    # Consistency assessment (NOT safety validation)
    consistency_level: str
    consistency_justification: str
    
    # Detailed breakdowns
    sfs_details: Dict
    brr_details: Dict
    ecc_details: Dict
    urs_details: Dict
    
    # Anomaly summary
    detected_anomalies: List[Dict]
    unsupported_claims: List[str]
    
    # Causal analysis (for research-level diagnosis)
    causal_analysis: Optional[Dict] = None
    
    # Confidence ceiling metadata (GUARDRAILS)
    confidence_ceilings_applied: Dict = field(default_factory=lambda: {
        "sfs_ceiling": SFS_CEILING,
        "brr_ceiling": BRR_CEILING,
        "ecc_ceiling": ECC_CEILING,
        "esri_ceiling": ESRI_ABSOLUTE_CEILING,
    })
    was_esri_capped: bool = False
    original_esri: float = 0.0
    
    def to_dict(self) -> Dict:
        return {
            "incident_id": self.incident_id,
            "evaluation_timestamp": self.evaluation_timestamp,
            "scores": {
                "ESRI": round(self.esri, 4),
                "CCR": round(self.ccr, 3),
                "SFS": round(self.sfs, 3),
                "BRS": round(self.brr, 3),
                "BRR": round(self.brr, 3),
                "ECC": round(self.ecc, 3),
                "AGI": round(self.agi, 3),
                "URS": round(self.urs, 3),
                "EES": round(self.ees, 4),
            },
            "score_policy": {
                "primary_framework": "CCR + BRR + AGI + URS",
                "primary_formula": "EES = CCR * BRR * AGI * URS",
                "legacy_framework": "ESRI",
                "legacy_use": "comparison_only",
                "metric_aliases": {
                    "BRR": "BRS",
                    "BRS": "BRR",
                    "CCR": "SFS",
                },
                "urs_status": (
                    "fallback"
                    if self.urs_details.get("fallback_reason")
                    else "top_n"
                ),
                "note": (
                    "ESRI is retained only for historical comparison and migration continuity. "
                    "Primary method defense should use the upgraded evaluation framework."
                ),
            },
            "consistency_level": self.consistency_level,
            "consistency_justification": self.consistency_justification,
            "details": {
                "constraint_correctness": self.sfs_details,
                "behavior_reproduction": self.brr_details,
                "evidence_consistency": self.ecc_details,
                "uncertainty_robustness": self.urs_details,
                "legacy_aliases": {
                    "scenario_fidelity": self.sfs_details,
                },
            },
            "detected_anomalies": self.detected_anomalies,
            "unsupported_claims": self.unsupported_claims,
            "causal_analysis": self.causal_analysis,
            # Include confidence ceiling metadata for transparency
            "confidence_controls": {
                "ceilings_applied": self.confidence_ceilings_applied,
                "was_esri_capped": self.was_esri_capped,
                "original_esri": round(self.original_esri, 4) if self.was_esri_capped else None,
                "rationale": (
                    "ESRI is capped at 0.85 maximum because FAA sighting reports are "
                    "observational/non-authoritative and PX4 simulation is a proxy model."
                ),
            },
        }
    
    def to_row(self) -> Dict:
        """Convert to flat row for Excel export."""
        return {
            "Incident ID": self.incident_id,
            "Timestamp": self.evaluation_timestamp,
            "ESRI": round(self.esri, 4),
            "CCR": round(self.ccr, 3),
            "BRS": round(self.brr, 3),
            "SFS": round(self.sfs, 3),
            "BRR": round(self.brr, 3),
            "ECC": round(self.ecc, 3),
            "AGI": round(self.agi, 3),
            "URS": round(self.urs, 3),
            "EES": round(self.ees, 4),
            "Consistency Level": self.consistency_level,
            "Anomaly Count": len(self.detected_anomalies),
            "Unsupported Claims": len(self.unsupported_claims),
            "ESRI Capped": "Yes" if self.was_esri_capped else "No",
            "SFS Confidence": self.sfs_details.get("confidence", ""),
            "BRR Confidence": self.brr_details.get("confidence", ""),
            "ECC Confidence": self.ecc_details.get("confidence", ""),
        }


class CaseEvaluator:
    """
    Orchestrates evaluation of a single incident case using all 4 metrics.
    
    Usage:
        evaluator = CaseEvaluator()
        result = evaluator.evaluate(
            faa_report=incident,
            px4_config=config,
            telemetry=telemetry_data,
            safety_report=safety_analysis,
        )
    """
    
    def __init__(self):
        self.ccr_scorer = ConstraintCorrectnessEvaluator()
        self.sfs_scorer = ScenarioFidelityScorer()  # Kept for backward compatibility if needed elsewhere
        self.brr_validator = BehaviorValidator()
        self.ecc_checker = EvidenceConsistencyChecker()
        self.urs_evaluator = UncertaintyRobustnessEvaluator()
        self.esri_calculator = ESRICalculator()
        self.causal_analyzer = SubsystemCausalAnalyzer()
        
        logger.debug("CaseEvaluator initialized with all metric components")
    
    def evaluate(
        self,
        faa_report: Dict,
        px4_config: Dict,
        telemetry: List[Dict],
        safety_report: Dict,
        telemetry_stats: Optional[Dict] = None,
    ) -> CaseEvaluationResult:
        """
        Evaluate a complete incident case.
        
        Args:
            faa_report: Original FAA incident data
            px4_config: LLM-generated PX4 configuration
            telemetry: Raw telemetry data points
            safety_report: Generated safety report
            telemetry_stats: Pre-computed telemetry statistics (optional)
            
        Returns:
            CaseEvaluationResult with all metrics and trust assessment
            
        SAFETY GUARDRAILS:
        - Component scores are capped at evidence-based ceilings
        - ESRI is capped at 0.85 maximum (FAA source is non-authoritative)
        - All capping is logged and recorded in result
        """
        incident_id = faa_report.get("incident_id", px4_config.get("faa_source", {}).get("incident_id", "unknown"))
        
        logger.info(f"Evaluating case: {incident_id}")
        
        # 1. Compute CCR (Constraint Correctness Rate)
        ccr_result = self.ccr_scorer.evaluate(faa_report, px4_config)

        # Apply legacy SFS ceiling to primary LLM translation score (CCR)
        raw_ccr = ccr_result.score
        capped_ccr = min(raw_ccr, SFS_CEILING)
        if capped_ccr < raw_ccr:
            logger.info(f"CCR capped: {raw_ccr:.3f} → {capped_ccr:.3f} (ceiling: {SFS_CEILING})")
            ccr_result.score = capped_ccr
        
        # 2. Compute BRR (Behavior Reproduction Rate)
        fault_type = px4_config.get("fault_injection", {}).get("fault_type", "")
        brr_result = self.brr_validator.evaluate(telemetry, fault_type, telemetry_stats)
        
        # Apply BRR ceiling (physics simulation accuracy)
        raw_brr = brr_result.score
        capped_brr = min(raw_brr, BRR_CEILING)
        if capped_brr < raw_brr:
            logger.info(f"BRR capped: {raw_brr:.3f} → {capped_brr:.3f} (ceiling: {BRR_CEILING})")
            brr_result.score = capped_brr
        
        # 3. Compute ECC (Evidence-Conclusion Consistency)
        detected_anomalies_dicts = [a.to_dict() for a in brr_result.detected_anomalies]
        stats = telemetry_stats or self.brr_validator._compute_telemetry_stats(telemetry)
        ecc_result = self.ecc_checker.evaluate(
            safety_report,
            detected_anomalies_dicts,
            stats
        )
        agi_score = ecc_result.agi_score

        # 4. Compute URS (Uncertainty Robustness Score) scaffold.
        alternative_configs = self._extract_alternative_configs(px4_config, safety_report)
        base_verdict = (
            safety_report.get("safety_level")
            or safety_report.get("hazard_level")
            or safety_report.get("risk_level")
        )
        urs_result = self.urs_evaluator.evaluate(
            primary_config=px4_config,
            alternative_configs=alternative_configs,
            base_verdict=base_verdict,
        )
        
        # Apply ECC ceiling (can be perfect if well-supported)
        raw_ecc = ecc_result.score
        capped_ecc = min(raw_ecc, ECC_CEILING)
        if capped_ecc < raw_ecc:
            logger.info(f"ECC capped: {raw_ecc:.3f} → {capped_ecc:.3f} (ceiling: {ECC_CEILING})")
            ecc_result.score = capped_ecc
        
        # 5. Perform Causal Subsystem Analysis (for research-level diagnosis)
        # This determines which subsystem failed FIRST and builds a causal chain
        causal_result = self.causal_analyzer.analyze(detected_anomalies_dicts)
        causal_analysis_dict = causal_result.to_dict()
        
        logger.info(
            f"Causal analysis: primary={causal_result.primary_failure_subsystem}, "
            f"confidence={causal_result.confidence:.2f}, conclusive={causal_result.is_conclusive}"
        )
        
        # 6. Compute ESRI (Executable Safety Reliability Index)
        # NOTE: ESRI calculator currently expects SFS key. We provide CCR as SFS alias
        # for backward compatibility while transitioning output nomenclature.
        esri_result = self.esri_calculator.calculate(
            {
                "SFS": ccr_result.score,
                "CCR": ccr_result.score,
                "confidence": ccr_result.confidence,
                "assessments": [a.to_dict() for a in ccr_result.assessments],
            },
            brr_result.to_dict(),
            ecc_result.to_dict(),
            incident_id
        )
        
        # =====================================================================
        # CRITICAL GUARDRAIL: Apply absolute ESRI ceiling
        # SAFETY RATIONALE: FAA sighting reports are observational and 
        # non-authoritative. PX4 simulation is a proxy model, not a 
        # reconstruction of the actual incident. ESRI must NEVER exceed 0.85.
        # =====================================================================
        raw_esri = esri_result.esri
        capped_esri = min(raw_esri, ESRI_ABSOLUTE_CEILING)
        was_capped = abs(capped_esri - raw_esri) > 0.001
        
        if was_capped:
            logger.warning(
                f"ESRI CAPPED BY GUARDRAIL: {raw_esri:.3f} → {capped_esri:.3f} "
                f"(ceiling: {ESRI_ABSOLUTE_CEILING}) - FAA source is non-authoritative"
            )
            esri_result.esri = capped_esri
            # Adjust consistency level if needed
            if esri_result.consistency_level == "HIGH" and capped_esri < 0.7:
                esri_result.consistency_level = "MEDIUM"
                esri_result.consistency_justification = (
                    f"ESRI capped from {raw_esri:.3f} to {capped_esri:.3f} by guardrail. "
                    + esri_result.consistency_justification
                )

        # Primary composite metric for upgraded framework.
        ees = ccr_result.score * brr_result.score * agi_score * urs_result.score
        
        # Build result with ceiling metadata
        result = CaseEvaluationResult(
            incident_id=incident_id,
            evaluation_timestamp=datetime.now().isoformat(),
            ccr=ccr_result.score,
            sfs=ccr_result.score,
            brr=brr_result.score,
            ecc=ecc_result.score,
            agi=agi_score,
            urs=urs_result.score,
            ees=ees,
            esri=esri_result.esri,
            consistency_level=esri_result.consistency_level,
            consistency_justification=esri_result.consistency_justification,
            sfs_details=ccr_result.to_dict(),
            brr_details=brr_result.to_dict(),
            ecc_details=ecc_result.to_dict(),
            urs_details=urs_result.to_dict(),
            detected_anomalies=detected_anomalies_dicts,
            unsupported_claims=ecc_result.unsupported_claims,
            causal_analysis=causal_analysis_dict,
            # Guardrail metadata
            was_esri_capped=was_capped,
            original_esri=raw_esri if was_capped else 0.0,
        )
        
        logger.info(
            f"Case evaluation complete: EES={result.ees * 100:.1f}%, ESRI={result.esri * 100:.1f}% "
            f"(CCR={result.ccr * 100:.0f}%, BRR={result.brr * 100:.0f}%, AGI={result.agi * 100:.0f}%, URS={result.urs * 100:.0f}%) "
            f"[{result.consistency_level}]{' [CAPPED]' if was_capped else ''}"
        )
        
        return result

    def _extract_alternative_configs(self, px4_config: Dict, safety_report: Dict) -> List[Dict]:
        """Collect top-N feasible alternatives if upstream components provided them."""
        for key in (
            "alternative_configs",
            "top_n_alternatives",
            "candidate_configs",
            "n_best_configs",
        ):
            value = px4_config.get(key)
            if isinstance(value, list):
                return [v for v in value if isinstance(v, dict)]

        for key in (
            "alternative_configs",
            "top_n_alternatives",
            "candidate_configs",
            "n_best_configs",
        ):
            value = safety_report.get(key)
            if isinstance(value, list):
                return [v for v in value if isinstance(v, dict)]

        return []
    
    def export_to_json(self, result: CaseEvaluationResult, output_path: Path) -> Path:
        """Export evaluation result to JSON file."""
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result.to_dict(), f, indent=2)
        
        logger.info(f"Evaluation exported to: {output_path}")
        return output_path


class EvaluationExcelExporter:
    """
    Exports evaluation results to Excel format.
    
    Generates both per-incident and aggregate reports.
    """
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_incident(
        self,
        result: CaseEvaluationResult,
        output_path: Optional[Path] = None
    ) -> Path:
        """
        Export single incident evaluation to Excel with comprehensive detail.
        
        Creates multiple sheets:
        - Summary: Scores, trust level, causal analysis overview
        - Anomalies: Full anomaly table with temporal info
        - Causal Analysis: Detailed subsystem breakdown
        - SFS Details: Dimension scores and keywords
        - ECC Claims: All verified claims with evidence
        """
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel export")
            return None
        
        output_path = output_path or (self.output_dir / f"evaluation_{result.incident_id}.xlsx")
        
        wb = Workbook()
        
        # Common styles
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        section_font = Font(bold=True, size=12, color="1A365D")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Trust level colors
        trust_fills = {
            "HIGH": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
            "LOW": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
        }
        
        # Severity colors
        severity_fills = {
            "CRITICAL": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
            "HIGH": PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid"),
            "LOW": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
        }
        
        # =====================================================================
        # SHEET 1: SUMMARY
        # =====================================================================
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary.merge_cells("A1:E1")
        ws_summary["A1"] = f"AeroGuardian Evaluation Report"
        ws_summary["A1"].font = Font(bold=True, size=16, color="1A365D")
        
        ws_summary.merge_cells("A2:E2")
        ws_summary["A2"] = f"Incident: {result.incident_id}"
        ws_summary["A2"].font = Font(size=12)
        
        ws_summary["A3"] = f"Generated: {result.evaluation_timestamp}"
        ws_summary["A3"].font = Font(size=10, italic=True)
        
        # Scores section
        row = 5
        ws_summary[f"A{row}"] = "EVALUATION SCORES"
        ws_summary[f"A{row}"].font = section_font
        
        row += 1
        headers = ["Metric", "Score", "Confidence", "Interpretation"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # ESRI row
        row += 1
        ws_summary.cell(row=row, column=1, value="ESRI (Internal Consistency Score)").border = thin_border
        ws_summary.cell(row=row, column=2, value=round(result.esri, 4)).border = thin_border
        consistency_cell = ws_summary.cell(row=row, column=3, value=result.consistency_level)
        consistency_cell.fill = trust_fills.get(result.consistency_level, trust_fills["LOW"])
        consistency_cell.font = Font(bold=True, color="FFFFFF")
        consistency_cell.border = thin_border
        ws_summary.cell(row=row, column=4, value=result.consistency_justification).border = thin_border
        
        # SFS row
        row += 1
        ws_summary.cell(row=row, column=1, value="SFS (Scenario Fidelity)").border = thin_border
        ws_summary.cell(row=row, column=2, value=round(result.sfs, 3)).border = thin_border
        ws_summary.cell(row=row, column=3, value=result.sfs_details.get("confidence", "")).border = thin_border
        fault_type = result.sfs_details.get("matched_fault_type", "unknown")
        ws_summary.cell(row=row, column=4, value=f"Matched fault: {fault_type}").border = thin_border
        
        # BRR row
        row += 1
        ws_summary.cell(row=row, column=1, value="BRR (Behavior Reproduction)").border = thin_border
        ws_summary.cell(row=row, column=2, value=round(result.brr, 3)).border = thin_border
        ws_summary.cell(row=row, column=3, value=result.brr_details.get("confidence", "")).border = thin_border
        ws_summary.cell(row=row, column=4, value=f"{len(result.detected_anomalies)} anomalies, {result.brr_details.get('data_points_analyzed', 0)} data points").border = thin_border
        
        # ECC row
        row += 1
        ws_summary.cell(row=row, column=1, value="ECC (Evidence Consistency)").border = thin_border
        ws_summary.cell(row=row, column=2, value=round(result.ecc, 3)).border = thin_border
        ws_summary.cell(row=row, column=3, value=result.ecc_details.get("confidence", "")).border = thin_border
        supported = result.ecc_details.get("supported_claims", 0)
        total = result.ecc_details.get("total_claims", 0)
        ws_summary.cell(row=row, column=4, value=f"{supported}/{total} claims supported").border = thin_border

        # AGI row
        row += 1
        ws_summary.cell(row=row, column=1, value="AGI (Actionability & Grounding)").border = thin_border
        ws_summary.cell(row=row, column=2, value=round(result.agi, 3)).border = thin_border
        ws_summary.cell(row=row, column=3, value=result.ecc_details.get("confidence", "")).border = thin_border
        agi_claim_count = result.ecc_details.get("agi_summary", {}).get("claim_count", 0)
        ws_summary.cell(row=row, column=4, value=f"{agi_claim_count} recommendation/constraint claims scored").border = thin_border

        # URS row
        row += 1
        ws_summary.cell(row=row, column=1, value="URS (Uncertainty Robustness)").border = thin_border
        ws_summary.cell(row=row, column=2, value=round(result.urs, 3)).border = thin_border
        ws_summary.cell(row=row, column=3, value=result.urs_details.get("confidence", "")).border = thin_border
        fallback_reason = result.urs_details.get("fallback_reason")
        ws_summary.cell(
            row=row,
            column=4,
            value=fallback_reason or f"Top-N alternatives: {result.urs_details.get('alternative_count', 0)}",
        ).border = thin_border

        # EES row
        row += 1
        ws_summary.cell(row=row, column=1, value="EES (Primary Composite)").border = thin_border
        ws_summary.cell(row=row, column=2, value=round(result.ees, 4)).border = thin_border
        ws_summary.cell(row=row, column=3, value="PRIMARY").border = thin_border
        ws_summary.cell(row=row, column=4, value="EES = CCR * BRS * AGI * URS").border = thin_border
        
        # Causal Analysis Summary (if available)
        if result.causal_analysis:
            row += 2
            ws_summary[f"A{row}"] = "CAUSAL ANALYSIS SUMMARY"
            ws_summary[f"A{row}"].font = section_font
            
            row += 1
            ca = result.causal_analysis
            primary = ca.get("primary_failure_subsystem", "undetermined")
            confidence = ca.get("confidence", 0)
            chain = ca.get("causal_chain", [])
            plausibility = ca.get("chain_plausibility", "unknown")
            
            ws_summary.cell(row=row, column=1, value="Root Cause Subsystem:").font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=primary.upper())
            ws_summary.cell(row=row, column=3, value=f"{confidence:.0%} confidence")
            
            row += 1
            ws_summary.cell(row=row, column=1, value="Causal Chain:").font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=" -> ".join(chain) if chain else "N/A")
            
            row += 1
            ws_summary.cell(row=row, column=1, value="Physics Plausibility:").font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=plausibility)
            if plausibility == "implausible":
                ws_summary.cell(row=row, column=2).font = Font(color="E74C3C")
            
            # Warnings
            warnings = ca.get("warnings", [])
            if warnings:
                row += 1
                ws_summary.cell(row=row, column=1, value="Warnings:").font = Font(bold=True)
                for i, warning in enumerate(warnings[:3]):
                    ws_summary.cell(row=row, column=2+i, value=warning).font = Font(color="E74C3C", italic=True)
        
        # Column widths
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 18
        ws_summary.column_dimensions["D"].width = 60
        
        # =====================================================================
        # SHEET 2: ANOMALIES
        # =====================================================================
        ws_anomalies = wb.create_sheet("Anomalies")
        
        ws_anomalies["A1"] = "DETECTED ANOMALIES"
        ws_anomalies["A1"].font = section_font
        
        headers = ["Type", "Severity", "Measured", "Threshold", "Subsystem", "First Detected (sec)", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws_anomalies.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for i, anomaly in enumerate(result.detected_anomalies, start=3):
            ws_anomalies.cell(row=i, column=1, value=anomaly.get("type", "")).border = thin_border
            
            severity = anomaly.get("severity", "")
            sev_cell = ws_anomalies.cell(row=i, column=2, value=severity)
            sev_cell.fill = severity_fills.get(severity, severity_fills["LOW"])
            sev_cell.font = Font(bold=True, color="FFFFFF")
            sev_cell.border = thin_border
            
            ws_anomalies.cell(row=i, column=3, value=round(anomaly.get("measured", 0), 3)).border = thin_border
            ws_anomalies.cell(row=i, column=4, value=anomaly.get("threshold", 0)).border = thin_border
            ws_anomalies.cell(row=i, column=5, value=anomaly.get("subsystem", "")).border = thin_border
            ws_anomalies.cell(row=i, column=6, value=anomaly.get("first_detected_sec", 0)).border = thin_border
            ws_anomalies.cell(row=i, column=7, value=anomaly.get("description", "")).border = thin_border
        
        # Column widths
        for col, width in enumerate([18, 12, 12, 12, 12, 18, 60], 1):
            ws_anomalies.column_dimensions[get_column_letter(col)].width = width
        
        # =====================================================================
        # SHEET 3: CAUSAL ANALYSIS (if available)
        # =====================================================================
        if result.causal_analysis:
            ws_causal = wb.create_sheet("Causal Analysis")
            ca = result.causal_analysis
            
            ws_causal["A1"] = "CAUSAL ANALYSIS DETAIL"
            ws_causal["A1"].font = section_font
            
            row = 3
            # Primary info
            ws_causal.cell(row=row, column=1, value="Primary Failure Subsystem:").font = Font(bold=True)
            ws_causal.cell(row=row, column=2, value=ca.get("primary_failure_subsystem", "undetermined").upper())
            ws_causal.cell(row=row, column=3, value=f"{ca.get('confidence', 0):.0%} confidence")
            
            row += 1
            ws_causal.cell(row=row, column=1, value="Is Conclusive:").font = Font(bold=True)
            ws_causal.cell(row=row, column=2, value="Yes" if ca.get("is_conclusive") else "No")
            
            row += 1
            ws_causal.cell(row=row, column=1, value="Causal Chain:").font = Font(bold=True)
            chain = ca.get("causal_chain", [])
            ws_causal.cell(row=row, column=2, value=" -> ".join(chain) if chain else "N/A")
            
            row += 1
            ws_causal.cell(row=row, column=1, value="Chain Plausibility:").font = Font(bold=True)
            ws_causal.cell(row=row, column=2, value=ca.get("chain_plausibility", "unknown"))
            
            # Diagnosis reasoning
            row += 2
            ws_causal.cell(row=row, column=1, value="Diagnosis Reasoning:").font = Font(bold=True)
            row += 1
            reasoning = ca.get("diagnosis_reasoning", "")
            ws_causal.merge_cells(f"A{row}:D{row}")
            ws_causal.cell(row=row, column=1, value=reasoning).alignment = Alignment(wrap_text=True)
            
            # Subsystem Evidence Table
            row += 2
            ws_causal.cell(row=row, column=1, value="SUBSYSTEM EVIDENCE").font = section_font
            
            row += 1
            headers = ["Subsystem", "Status", "Confidence", "First Anomaly (sec)", "Anomalies"]
            for col, header in enumerate(headers, 1):
                cell = ws_causal.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            subsystem_evidence = ca.get("subsystem_evidence", {})
            for subsystem, data in subsystem_evidence.items():
                row += 1
                ws_causal.cell(row=row, column=1, value=subsystem).border = thin_border
                
                status = data.get("status", "")
                status_cell = ws_causal.cell(row=row, column=2, value=status)
                if status == "CONFIRMED":
                    status_cell.font = Font(bold=True, color="27AE60")
                elif status == "LIKELY":
                    status_cell.font = Font(color="F39C12")
                status_cell.border = thin_border
                
                ws_causal.cell(row=row, column=3, value=f"{data.get('confidence', 0):.0%}").border = thin_border
                
                first_time = data.get("first_anomaly_time_sec")
                ws_causal.cell(row=row, column=4, value=first_time if first_time is not None else "N/A").border = thin_border
                
                anomalies_list = data.get("anomalies", [])
                ws_causal.cell(row=row, column=5, value=", ".join(anomalies_list[:5])).border = thin_border
            
            # Warnings
            warnings = ca.get("warnings", [])
            if warnings:
                row += 2
                ws_causal.cell(row=row, column=1, value="WARNINGS").font = Font(bold=True, color="E74C3C")
                for warning in warnings:
                    row += 1
                    ws_causal.cell(row=row, column=1, value=warning).font = Font(color="E74C3C")
            
            # Column widths
            for col, width in enumerate([20, 15, 12, 18, 50], 1):
                ws_causal.column_dimensions[get_column_letter(col)].width = width
        
        # =====================================================================
        # SHEET 4: SFS DETAILS
        # =====================================================================
        ws_sfs = wb.create_sheet("SFS Details")
        
        ws_sfs["A1"] = "SCENARIO FIDELITY SCORE BREAKDOWN"
        ws_sfs["A1"].font = section_font
        
        row = 3
        # Dimension scores
        headers = ["Dimension", "Score", "Weight", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws_sfs.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        dimension_descriptions = {
            "fault_type_match": "How well the simulated fault matches FAA narrative",
            "trigger_condition_match": "Match of trigger conditions (time, environmental factors)",
            "environmental_match": "Weather and environmental condition alignment",
            "temporal_consistency": "Timeline consistency between simulation and report",
            "parameter_completeness": "Completeness of simulation parameters",
        }
        
        dimensions = result.sfs_details.get("dimension_scores", {})
        for dim_name, dim_score in dimensions.items():
            row += 1
            ws_sfs.cell(row=row, column=1, value=dim_name.replace("_", " ").title()).border = thin_border
            ws_sfs.cell(row=row, column=2, value=round(dim_score, 3)).border = thin_border
            ws_sfs.cell(row=row, column=3, value="20%").border = thin_border  # Equal weights
            ws_sfs.cell(row=row, column=4, value=dimension_descriptions.get(dim_name, "")).border = thin_border
        
        # Other SFS details
        row += 2
        ws_sfs.cell(row=row, column=1, value="Matched Fault Type:").font = Font(bold=True)
        ws_sfs.cell(row=row, column=2, value=result.sfs_details.get("matched_fault_type", "unknown"))
        
        row += 1
        ws_sfs.cell(row=row, column=1, value="Extracted Keywords:").font = Font(bold=True)
        keywords = result.sfs_details.get("extracted_keywords", [])
        ws_sfs.cell(row=row, column=2, value=", ".join(keywords) if keywords else "None")
        
        row += 1
        ws_sfs.cell(row=row, column=1, value="Missing Parameters:").font = Font(bold=True)
        missing = result.sfs_details.get("missing_parameters", [])
        ws_sfs.cell(row=row, column=2, value=", ".join(missing) if missing else "None")
        
        # Column widths
        ws_sfs.column_dimensions["A"].width = 25
        ws_sfs.column_dimensions["B"].width = 12
        ws_sfs.column_dimensions["C"].width = 10
        ws_sfs.column_dimensions["D"].width = 55
        
        # =====================================================================
        # SHEET 5: ECC CLAIMS
        # =====================================================================
        ws_ecc = wb.create_sheet("ECC Claims")
        
        ws_ecc["A1"] = "EVIDENCE-CONCLUSION CONSISTENCY ANALYSIS"
        ws_ecc["A1"].font = section_font
        
        # Summary
        ws_ecc["A3"] = f"Total Claims: {result.ecc_details.get('total_claims', 0)}"
        ws_ecc["B3"] = f"Supported: {result.ecc_details.get('supported_claims', 0)}"
        ws_ecc["C3"] = f"Evidence Strength: {result.ecc_details.get('evidence_strength', 0):.2f}"
        
        row = 5
        headers = ["Claim Type", "Claim Text", "Supported", "Confidence", "Evidence"]
        for col, header in enumerate(headers, 1):
            cell = ws_ecc.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        verified_claims = result.ecc_details.get("verified_claims", [])
        for claim in verified_claims[:50]:  # Limit to 50 claims
            row += 1
            ws_ecc.cell(row=row, column=1, value=claim.get("claim_type", "")).border = thin_border
            ws_ecc.cell(row=row, column=2, value=claim.get("claim_text", "")[:80]).border = thin_border
            
            supported = claim.get("is_supported", False)
            supp_cell = ws_ecc.cell(row=row, column=3, value="Yes" if supported else "No")
            supp_cell.font = Font(bold=True, color="27AE60" if supported else "E74C3C")
            supp_cell.border = thin_border
            
            ws_ecc.cell(row=row, column=4, value=f"{claim.get('confidence', 0):.0%}").border = thin_border
            
            evidence = claim.get("supporting_evidence", [])
            ws_ecc.cell(row=row, column=5, value="; ".join(evidence[:3])).border = thin_border
        
        # Unsupported claims (if any)
        if result.unsupported_claims:
            row += 2
            ws_ecc.cell(row=row, column=1, value="UNSUPPORTED CLAIMS").font = Font(bold=True, color="E74C3C")
            for claim in result.unsupported_claims[:10]:
                row += 1
                ws_ecc.cell(row=row, column=1, value=claim).font = Font(color="E74C3C")
        
        # Column widths
        for col, width in enumerate([15, 45, 12, 12, 50], 1):
            ws_ecc.column_dimensions[get_column_letter(col)].width = width
        
        # =====================================================================
        # SAVE WORKBOOK
        # =====================================================================
        try:
            wb.save(output_path)
            logger.info(f"Evaluation Excel exported: {output_path} (5 sheets)")
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            raise
        
        return output_path
    
    def export_aggregate(
        self,
        results: List[CaseEvaluationResult],
        output_path: Optional[Path] = None
    ) -> Path:
        """Export aggregate evaluation across multiple incidents."""
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel export")
            return None
        
        output_path = output_path or (self.output_dir / "aggregate_evaluation.xlsx")
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary.merge_cells("A1:G1")
        ws_summary["A1"] = "AeroGuardian Aggregate Evaluation Report"
        ws_summary["A1"].font = Font(bold=True, size=14)
        
        ws_summary["A3"] = f"Total Incidents: {len(results)}"
        ws_summary["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Aggregate stats
        if results:
            esri_scores = [r.esri for r in results]
            ws_summary["A6"] = "Aggregate Statistics"
            ws_summary["A6"].font = Font(bold=True)
            ws_summary["A7"] = f"Average ESRI: {sum(esri_scores)/len(esri_scores):.4f}"
            ws_summary["A8"] = f"Min ESRI: {min(esri_scores):.4f}"
            ws_summary["A9"] = f"Max ESRI: {max(esri_scores):.4f}"
            
            consistency_counts = {
                "HIGH": sum(1 for r in results if r.consistency_level == "HIGH"),
                "MEDIUM": sum(1 for r in results if r.consistency_level == "MEDIUM"),
                "LOW": sum(1 for r in results if r.consistency_level == "LOW"),
            }
            ws_summary["A11"] = "Consistency Distribution"
            ws_summary["A11"].font = Font(bold=True)
            ws_summary["A12"] = f"HIGH: {consistency_counts['HIGH']} ({100*consistency_counts['HIGH']/len(results):.1f}%)"
            ws_summary["A13"] = f"MEDIUM: {consistency_counts['MEDIUM']} ({100*consistency_counts['MEDIUM']/len(results):.1f}%)"
            ws_summary["A14"] = f"LOW: {consistency_counts['LOW']} ({100*consistency_counts['LOW']/len(results):.1f}%)"
        
        # Sheet 2: Per-Incident Details
        ws_details = wb.create_sheet("Per-Incident Details")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        headers = list(results[0].to_row().keys()) if results else []
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, result in enumerate(results, 2):
            row_data = result.to_row()
            for col_idx, header in enumerate(headers, 1):
                ws_details.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
        
        # Adjust column widths
        for col_idx, header in enumerate(headers, 1):
            ws_details.column_dimensions[get_column_letter(col_idx)].width = max(len(str(header)) + 2, 12)
        
        wb.save(output_path)
        logger.info(f"Aggregate evaluation Excel exported: {output_path}")
        return output_path


# =============================================================================
# SINGLETON ACCESSOR
# =============================================================================

_evaluator: Optional[CaseEvaluator] = None

def get_case_evaluator() -> CaseEvaluator:
    """Get singleton case evaluator instance."""
    global _evaluator
    if _evaluator is None:
        _evaluator = CaseEvaluator()
    return _evaluator


def get_column_letter(col_idx: int) -> str:
    """Convert column index to Excel letter (1=A, 2=B, etc.)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result
